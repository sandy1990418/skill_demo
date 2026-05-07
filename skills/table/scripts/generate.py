"""
Generate a styled table as PNG or Excel from JSON/CSV data.

Usage:
    python generate.py --data data.json --output table.png
    python generate.py --data data.json --output table.xlsx
    python generate.py --data data.json --output table.csv

data.json format: list-of-dicts, e.g. [{"Name": "A", "Value": 100}, ...]
"""

import argparse
import json
import os
import sys
from pathlib import Path


def load_data(path: str):
    import pandas as pd

    p = Path(path)
    if p.suffix == ".csv":
        return pd.read_csv(p)
    data = json.loads(p.read_text(encoding="utf-8"))
    return pd.DataFrame(data)


def to_png(df, output_path: str, title: str | None):
    from matplotlib.backends.backend_agg import FigureCanvasAgg
    from matplotlib.figure import Figure

    n_rows, n_cols = df.shape
    fig = Figure(figsize=(max(8, n_cols * 2), n_rows * 0.5 + 1.5))
    FigureCanvasAgg(fig)
    ax = fig.add_subplot(111)
    ax.axis("off")

    if title:
        ax.set_title(title, fontsize=13, fontweight="bold", pad=10)

    tbl = ax.table(
        cellText=df.values,
        colLabels=df.columns,
        cellLoc="center",
        loc="center",
    )
    tbl.auto_set_font_size(False)
    tbl.set_fontsize(11)
    tbl.scale(1, 1.6)

    for j in range(n_cols):
        tbl[0, j].set_facecolor("#1E2761")
        tbl[0, j].set_text_props(color="white", fontweight="bold")

    for i in range(1, n_rows + 1):
        for j in range(n_cols):
            tbl[i, j].set_facecolor("#F5F5F5" if i % 2 == 0 else "white")

    fig.tight_layout()
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(output_path, dpi=150, bbox_inches="tight")


def to_xlsx(df, output_path: str):
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    with __import__("pandas").ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Data")
        ws = writer.sheets["Data"]

        header_fill = PatternFill("solid", fgColor="1E2761")
        header_font = Font(color="FFFFFF", bold=True)
        alt_fill = PatternFill("solid", fgColor="F5F5F5")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
            for cell in row:
                if i % 2 == 0:
                    cell.fill = alt_fill
                cell.alignment = Alignment(horizontal="center")

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 18


def main():
    parser = argparse.ArgumentParser(description="Generate a styled table from JSON/CSV")
    parser.add_argument("--data", required=True, help="Path to JSON or CSV data file")
    parser.add_argument("--output", default="temp/table.png", help="Output path (.png / .xlsx / .csv)")
    parser.add_argument("--title", help="Table title (PNG only)")
    args = parser.parse_args()

    output = os.environ.get("TABLE_OUTPUT", args.output)
    ext = Path(output).suffix.lower()

    print(f"[table] loading {args.data}", file=sys.stderr)
    df = load_data(args.data)
    print(f"[table] {df.shape[0]} rows × {df.shape[1]} cols → {output}", file=sys.stderr)

    if ext == ".png":
        to_png(df, output, args.title)
    elif ext in (".xlsx", ".xls"):
        to_xlsx(df, output)
    elif ext == ".csv":
        Path(output).parent.mkdir(parents=True, exist_ok=True)
        df.to_csv(output, index=False)
    else:
        print(f"[error] unsupported output format: {ext}", file=sys.stderr)
        sys.exit(1)

    print(f"[done] {output}", file=sys.stderr)
    print(output)


if __name__ == "__main__":
    main()
